#
import os, sys, xlwt
import config
from public.scheduler.baseScheduler import *
from libs.utils.debug import *
import openpyxl

subDirs = "/tmp1/temp/"

def saveExcel(path, structs, rows) :

	# 获取文件扩展名
	fileType = "xlsx"
	index = path.rindex(".")
	if index > 0:
		fileType = path[index + 1:]		
	else:
		# 文件名没有扩展名，默认为xlsx格式
		pass

	if fileType == "xls":
		# xls 格式
		wb = xlwt.Workbook()
		sh = wb.add_sheet("Sheet1")

		# 标题
		fields = structs.split(",")
		for i in range(0, len(fields)):
			v = fields[i]
			v = v.strip()
			sh.write(0, i, v)

		r = 1
		for row in rows:
			for i in range(0, len(row)):
				v = row[i]
				#v = v.strip()
				sh.write(r, i, v)
			r += 1

		wb.save(path)

	else:
		# xlsx 格式
		wb = openpyxl.Workbook()

		# 使用默认的工作簿
		ws = wb.create_sheet("Sheet1", index = 0)

		# 标题
		fields = structs.split(",")
		title = []
		for key in fields:
			key = key.strip()
			title.append(key)

		ws.append(title)
		for row in rows:
			ws.append(row)

		wb.save(path)	

	return

def createTempFile(ext):
	name = randomStr(10) + "." + ext
	url = subDirs + name
	return {
		"url" : url,
		"path" : config.FileUploadRootPath + url
	}

# 创建Excel文件
def CreateExcelFile(structs, rows):
	name 	= randomStr(10) + ".xlsx"
	url 	= subDirs + name
	path 	= config.FileUploadRootPath + url
	
	saveExcel(path, structs, rows)

	return {
		"url" : url,
		"path" : config.FileUploadRootPath + url
	}



# 每天晚上1点执行任务
@scheduler_func(hour = "1", minute = "0", second = "0")
def deleteTempFile():
	path = config.FileUploadRootPath + subDirs
	time = datetime.datetime.now()
	for parent, dirnames, filenames in os.walk(path):
		for filename in filenames:
			filepath = os.path.join(parent, filename)
			mtime = os.path.getmtime(filepath)
			mtime = datetime.datetime.fromtimestamp(mtime)
			delta = time - mtime
			#D("%s %ss" % (filepath, delta.seconds))
			if delta.seconds > 30 * 60:
				logD("remove temp file %s " % filepath)
				os.remove(filepath)

